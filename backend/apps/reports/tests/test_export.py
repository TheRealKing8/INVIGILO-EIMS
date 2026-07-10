"""Tests for the PDF / Excel / CSV renderers and the export pipeline."""
from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from reportlab.pdfgen import canvas
from rest_framework.test import APIClient

from apps.academic.models import Course, Department, Faculty, Program
from apps.exams.models import ExamPeriod, ExamSession
from apps.invigilators.models import InvigilatorProfile
from apps.reports.models import ReportExport
from apps.reports.services.csv import render_csv
from apps.reports.services.excel import render_workbook
from apps.reports.services.pdf import render_attendance_summary
from apps.rooms.models import Building, Room


User = get_user_model()
pytestmark = pytest.mark.django_db


@pytest.fixture
def cycle(db):  # type: ignore[no-untyped-def]
    f = Faculty.objects.create(code="RPT-F", name="F")
    d = Department.objects.create(faculty=f, code="RPT-D", name="D")
    p = Program.objects.create(department=d, code="RPT-P", name="P")
    course = Course.objects.create(program=p, code="RPT-C", title="C", credit_hours=3)
    b = Building.objects.create(code="RPT-B", name="B")
    room = Room.objects.create(building=b, code="RPT-R", capacity=50)
    period = ExamPeriod.objects.create(
        code="RPT-T1", name="Report Test Term", is_active=True,
        starts_on=date.today(), ends_on=date.today() + timedelta(days=30),
    )
    ExamSession.objects.create(
        period=period, course=course, room=room,
        starts_at=datetime(2026, 8, 1, 9, 0, tzinfo=timezone.utc),
        ends_at=datetime(2026, 8, 1, 11, 0, tzinfo=timezone.utc),
        capacity=50, registered=42, invigilators_required=2, status="scheduled",
    )
    return period


# ---------------------------------------------------------------------------
# Renderer unit tests
# ---------------------------------------------------------------------------
def test_pdf_renderer_returns_valid_pdf(cycle) -> None:  # type: ignore[no-untyped-def]
    body = render_attendance_summary(cycle)
    assert body[:4] == b"%PDF", "Output is not a valid PDF header"


def test_excel_renderer_returns_xlsx(cycle) -> None:  # type: ignore[no-untyped-def]
    body = render_workbook(cycle)
    # .xlsx is a zip — magic bytes are 'PK\x03\x04'.
    assert body[:2] == b"PK", "Output is not a valid xlsx header"
    # And openpyxl can read it back.
    wb = openpyxl.load_workbook(BytesIO(body))
    assert "Sessions" in wb.sheetnames
    assert "Invigilators" in wb.sheetnames
    assert "Allocations" in wb.sheetnames


def test_csv_renderer_handles_incidents() -> None:
    from apps.incidents.models import Incident

    Incident.objects.create(title="X", severity="low", status="open")
    body = render_csv("incidents").decode("utf-8")
    assert "id,title,severity,status" in body
    assert "X" in body


def test_csv_renderer_unknown_model_returns_placeholder() -> None:
    body = render_csv("nope")
    assert b"unknown" in body


# ---------------------------------------------------------------------------
# End-to-end export pipeline
# ---------------------------------------------------------------------------
def test_export_pdf_round_trip(
    client: APIClient, verified_user, grant_permission, cycle
) -> None:
    grant_permission(verified_user, "report.view", "report.export")
    client.force_authenticate(verified_user)
    create = client.post(
        reverse("reports:report-export-list"),
        {"title": "Attendance", "format": "pdf", "cycle_id": str(cycle.id)},
        format="json",
    )
    assert create.status_code == 201, create.json()
    body = create.json()
    assert body["format"] == "pdf"
    # Eager-mode Celery -> the file is attached.
    response = client.get(reverse("reports:report-export-download", args=[body["id"]]))
    assert response.status_code == 200
    file_body = b"".join(response.streaming_content)
    assert file_body[:4] == b"%PDF"


def test_export_excel_round_trip(
    client: APIClient, verified_user, grant_permission, cycle
) -> None:
    grant_permission(verified_user, "report.view", "report.export")
    client.force_authenticate(verified_user)
    create = client.post(
        reverse("reports:report-export-list"),
        {"title": "Workbook", "format": "excel", "cycle_id": str(cycle.id)},
        format="json",
    )
    assert create.status_code == 201, create.json()
    response = client.get(
        reverse("reports:report-export-download", args=[create.json()["id"]])
    )
    assert response.status_code == 200
    file_body = b"".join(response.streaming_content)
    assert file_body[:2] == b"PK"


def test_export_csv_round_trip(
    client: APIClient, verified_user, grant_permission
) -> None:
    grant_permission(verified_user, "report.view", "report.export")
    client.force_authenticate(verified_user)
    create = client.post(
        reverse("reports:report-export-list"),
        {
            "title": "Incidents",
            "format": "csv",
            "parameters": {"model": "incidents"},
        },
        format="json",
    )
    assert create.status_code == 201
    response = client.get(
        reverse("reports:report-export-download", args=[create.json()["id"]])
    )
    body = b"".join(response.streaming_content)
    assert b"id,title,severity,status" in body
