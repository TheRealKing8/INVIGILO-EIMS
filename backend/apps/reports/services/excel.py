"""Excel report renderers using openpyxl."""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.allocations.models import Allocation
from apps.exams.models import ExamSession
from apps.invigilators.models import InvigilatorProfile


_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def render_workbook(period, sessions: Iterable[ExamSession] | None = None) -> bytes:
    """Multi-sheet .xlsx: one sheet per domain.

    Sheets: Sessions, Invigilators, Allocations, Incidents.
    """
    buffer = io.BytesIO()
    wb = Workbook()

    # --- Sessions ---
    ws = wb.active
    ws.title = "Sessions"
    _header(ws, ["Course", "Title", "Starts", "Ends", "Room", "Capacity", "Registered", "Status"])
    qs = sessions or ExamSession.objects.filter(period=period).select_related("course", "room")
    for s in qs:
        ws.append(
            [
                s.course.code,
                s.course.title,
                s.starts_at.replace(tzinfo=None),
                s.ends_at.replace(tzinfo=None),
                s.room.code if s.room else "",
                s.capacity,
                s.registered,
                s.status,
            ]
        )

    # --- Invigilators ---
    inv_wb = wb.create_sheet("Invigilators")
    _header(inv_wb, ["Name", "Email", "Department", "Max / cycle", "Rating"])
    for p in InvigilatorProfile.objects.select_related("user", "primary_department"):
        inv_wb.append(
            [
                p.user.full_name,
                p.user.email,
                p.primary_department.code if p.primary_department else "",
                p.max_sessions_per_cycle,
                float(p.rating),
            ]
        )

    # --- Allocations ---
    alloc_wb = wb.create_sheet("Allocations")
    _header(alloc_wb, ["Course", "Starts", "Invigilator", "Role", "Status", "Room"])
    for a in Allocation.objects.filter(run__period=period).select_related(
        "session__course", "invigilator__user", "room"
    ):
        alloc_wb.append(
            [
                a.session.course.code,
                a.session.starts_at.replace(tzinfo=None),
                a.invigilator.user.full_name,
                a.role,
                a.status,
                a.room.code if a.room else "",
            ]
        )

    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["render_workbook"]
